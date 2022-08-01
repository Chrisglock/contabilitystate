from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def Establecer_Estado(Elaboracion: str, Vencimiento: str) -> str:
    
    Elaboracion_Partes: list(str) = Elaboracion.split("/")
    Vencimiento_Partes: list(str) = Vencimiento.split("/")

    Inicio: date = date(int(Elaboracion_Partes[2]), int(Elaboracion_Partes[1]), int(Elaboracion_Partes[0]))
    Actual: date = date.today()
    Final: date = date(int(Vencimiento_Partes[2]), int(Vencimiento_Partes[1]), int(Vencimiento_Partes[0]))

    if Actual > Inicio and Actual < Final:
        return "Vigente"
    
    elif Actual == Final:
        return "Caducado"
    
    else:
        return "Caducado"

class Producto:

    def __init__(self, Identificador, Nombre, Cantidad, Precio_Compra, Precio_Venta, Elaboracion, Vencimiento, Proveedor):

        self.Identificador: int = Identificador
        self.Nombre: str = Nombre
        self.Cantidad: int = Cantidad
        self.Precio_Compra: float = Precio_Compra
        self.Precio_Venta: float = Precio_Venta
        self.Elaboracion: str = Elaboracion
        self.Vencimiento: str = Vencimiento
        self.Estado: str = Establecer_Estado(Elaboracion, Vencimiento)
        self.Proveedor: str = Proveedor
    
    def Modificar_Nombre(self, Nombre: str):
        self.Nombre = Nombre
    
    def Modificar_Cantidad(self, Cantidad: int):
        self.Cantidad = Cantidad

    def Modificar_Precio_Compra(self, Precio_Compra: float):
        self.Precio_Compra = Precio_Compra

    def Modificar_Precio_Venta(self, Precio_Venta: float):
        self.Precio_Venta = Precio_Venta
    
    def Modificar_Elaboracion(self, Elaboracion: str):
        self.Elaboracion = Elaboracion
        self.Estado = Establecer_Estado(self.Elaboracion, self.Vencimiento)
    
    def Modificar_Vencimiento(self, Vencimiento: str):
        self.Vencimiento = Vencimiento
        self.Estado = Establecer_Estado(self.Elaboracion, self.Vencimiento)

    def Modificar_Proveedor(self, Proveedor: str):
        self.Proveedor = Proveedor
    
    def __str__(self):
        return "Producto [ Identificador: {} | Nombre: {} | Cantidad: {} | Precio Compra: {} | Precio Venta: {} | Elaboracion: {} | Vencimiento: {} | Estado: {} | Proveedor: {} ]".format(self.Identificador, self.Nombre, self.Cantidad, int(self.Precio_Compra), int(self.Precio_Venta), self.Elaboracion, self.Vencimiento, self.Estado, self.Proveedor)

def Abrir_Stock(Nombre_Archivo: str) -> list:

    Archivo = load_workbook(filename=Nombre_Archivo)
    Informacion = Archivo.active

    Productos = []

    for fila in range(1, Informacion.max_row+1):

        if fila > 1:

            Producto_Temporal = []

            for columna in range(1, Informacion.max_column+1):

                if columna == 6 or columna == 7:
                    Fecha = str(Informacion["{}{}".format(get_column_letter(columna), fila)].value)
                    Fecha = datetime.strptime(Fecha, '%Y-%m-%d %H:%M:%S')
                    Fecha = Fecha.strftime('%d/%m/%Y')
                    Producto_Temporal.append(Fecha)
                else:
                    Producto_Temporal.append(str(Informacion["{}{}".format(get_column_letter(columna), fila)].value))

            Producto_Final = Producto(Producto_Temporal[0], Producto_Temporal[1], Producto_Temporal[2], Producto_Temporal[3], Producto_Temporal[4], Producto_Temporal[5], Producto_Temporal[6], Producto_Temporal[7])#, Producto_Temporal[8]) 

            Productos.append(Producto_Final)
        
    return Productos

class Stock:

    def __init__(self):
        self.Productos: dict= {}
    
    def Agregar_Stock(self, Archivo):

        if self.Productos:
            self.Productos[len(self.Productos)] = Abrir_Stock(Archivo)

        else:
            self.Productos[0] = Abrir_Stock(Archivo)

